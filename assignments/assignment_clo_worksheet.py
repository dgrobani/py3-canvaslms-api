# https://openpyxl.readthedocs.io/
# https://automatetheboringstuff.com/chapter12/
# https://www.ablebits.com/office-addins-blog/2014/09/24/excel-drop-down-list/
# http://stackoverflow.com/questions/18595686/how-does-operator-itemgetter-and-sort-work-in-python

from canvas.core.courses import get_course_by_sis_id, validate_course
from canvas.core.io import get_cmi_clos_by_course, tada
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, colors, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from canvas.core.assignments import get_assignments


def assignment_clo_worksheet():

    courses = {
        '2016-2SU-01-NDNP-714-LEC-ONL-O1': ['Paulina', 'Van'],
        '2016SS-OAK-UGAOAK1-NURSG-160-LEC1-1': ['Paulina', 'Van', 'NABSN'],
        '2016-3FA-02-NABSN-170-LEC-SFP-01': ['Jenny', 'Zettler Rhodes'],
        '2016-3FA-01-NBSN-164-LEC-OAK-01': ['Erik', 'Carter'],
        '2016-3FA-01-NELMSN-566-LEC-SAC-01': ['Erik', 'Carter'],
        '2016-3FA-01-NBSN-108-LEC-OAK-01': ['Christine', 'Rey']
        }

    for course_sis_id in courses:

        template_file = load_workbook('assignment_clo_worksheet.xlsx')
        sheet = template_file.get_sheet_by_name(template_file.active.title)

        sheet.freeze_panes = 'B1'
        sheet.page_setup.fitToHeight = 1
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        sixteen_point = Font(size=16)
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        sheet.add_data_validation(dv)

        teacher_firstname = courses[course_sis_id][0]
        teacher_lastname = courses[course_sis_id][1]
        course = get_course_by_sis_id(course_sis_id)
        course_sis_info = validate_course(course)
        program, number, ctype, campus, section, term, session = \
            [course_sis_info[i] for i in ['program', 'number', 'type', 'campus', 'section', 'term', 'session']]
        filename = '{}-{}-{}-{}-{}-{}-{}-{}.xlsx'\
            .format(program, number, ctype, campus, section, term, session, teacher_lastname)

        # header
        sheet.cell(row=1, column=1).value = number
        sheet.cell(row=2, column=1).value = course_sis_id
        sheet.cell(row=3, column=1).value = course['name']
        sheet.cell(row=1, column=2).value = term
        sheet.cell(row=2, column=2).value = teacher_firstname + ' ' + teacher_lastname

        # assignments (graded only)
        assignments = get_assignments(course['id'])
        for row, assignment in enumerate(sorted(assignments, key=lambda a: "" if not a['due_at'] else a['due_at'])):

            if 'not_graded' in assignment['submission_types'] or not assignment['points_possible'] \
                    or ('omit_from_final_grade' in assignment and assignment['omit_from_final_grade']):
                continue

            sheet.cell(row=7+row, column=1).value = assignment['name']
            sheet.cell(row=7+row, column=1).hyperlink = assignment['html_url']
            sheet.cell(row=7+row, column=1).border = border
            sheet.cell(row=7+row, column=1).font = sixteen_point
            sheet.cell(row=7+row, column=1).font = Font(color=colors.BLUE)
            sheet.row_dimensions[7+row].height = 27

            # rubric yes/no
            sheet.cell(row=7+row, column=2).border = border
            sheet.cell(row=7+row, column=2).font = sixteen_point
            dv.add(sheet.cell(row=7+row, column=2))

            # improvement plan
            sheet.cell(row=7+row, column=3).border = border

            # plan complete yes/no
            sheet.cell(row=7+row, column=4).border = border
            sheet.cell(row=7+row, column=4).font = sixteen_point
            dv.add(sheet.cell(row=7+row, column=4))

        # clos
        max_clo_desc_len = 0
        # kludge for old sis id format
        program = program if len(courses[course_sis_id]) == 2 else courses[course_sis_id][2]
        clos = get_cmi_clos_by_course(program, course_sis_info['number'])
        for col, clo in enumerate(clos):
            sheet.cell(row=6, column=5+col).alignment = Alignment(vertical='top', wrapText=True)
            sheet.cell(row=6, column=5+col).value = '{}: {}'.format(clo['clo_title'], clo['clo_description'])
            sheet.cell(row=6, column=5+col).border = border
            sheet.cell(row=6, column=5+col).font = sixteen_point
            max_clo_desc_len = max(len(clo['clo_description']), max_clo_desc_len)

        # clo headers [styling merged cells doesn't work in openpyxl]
        last_column = 4 + len(clos)
        sheet.merge_cells(start_row=4, start_column=5, end_row=4, end_column=last_column)
        sheet.merge_cells(start_row=5, start_column=5, end_row=5, end_column=last_column)

        # clo column width & row height
        sheet.row_dimensions[6].height = max_clo_desc_len / 50 * 36
        for column in range(5, last_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 50

        # conditional formatting for x marks the spot
        clo_range = 'E7:{}{}'.format(get_column_letter(last_column), 6 + len(assignments))
        sheet.conditional_formatting\
            .add(clo_range, CellIsRule(operator='greaterThan', formula=['""'], fill=PatternFill(bgColor='70AD47')))
        for row in range(7, 7 + len(assignments)):
            for column in range(5, last_column + 1):
                sheet.cell(row=row, column=column).border = border
                sheet.cell(row=row, column=column).alignment = Alignment(horizontal="center", vertical="center")

        template_file.save(filename)


if __name__ == '__main__':
    assignment_clo_worksheet()
    tada()
